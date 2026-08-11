"""
Applicant Evaluation Spreadsheet generator.

Builds the recruiter-facing PERM applicant review workbook: YES/NO evaluation
columns for primary / alternative / special-skills requirements, an optional
auto-recommendation formula ("Send Questionnaire" / "Do Not Send") with
row-highlight conditional formatting, outreach tracking, and the evaluator's
conclusion section. Modeled on the firm's existing Workday review template.

Everything here is stateless: config in, .xlsx bytes out. Nothing persisted.
"""
from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Style constants (lifted from the reference template) ────────────────────
FILL_GREY_LIGHT = "D9D9D9"   # header row / paste zone
FILL_PRIMARY    = "5B9BD5"   # blue    — primary requirements banner
FILL_ALT        = "70AD47"   # green   — alternative requirements banner
FILL_SKILLS     = "A5A5A5"   # grey    — special skills banner
FILL_OUTREACH   = "44546A"   # slate   — outreach banner
FILL_RESPONSES  = "FFC000"   # gold    — questionnaire responses banner
FILL_CONCLUDE   = "ED7D31"   # orange  — evaluator's conclusion banner
FILL_HIGHLIGHT  = "FFFF00"   # yellow  — recommended-row highlight

MAX_SKILLS = 26          # sanity bound; template had 10
DEFAULT_ROWS = 1000      # applicant rows (template shipped 5000; 1000 keeps file small)

DISCLAIMER = (
    "Note: The spreadsheet automatically generates a recommendation based on "
    "the evaluation criteria. However, the evaluator is responsible for "
    "reviewing the results and verifying that the spreadsheet's calculations "
    "and recommendations are accurate."
)

SEND = "Send Questionnaire"
NOSEND = "Do Not Send"

_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _sanitize(text: str) -> str:
    return re.sub(r"[\\/*?:\[\]]", "-", (text or "").strip())[:80]


# Conditions of employment, not skills — excluded from the special-skills
# columns (the template's Evaluator Conclusion covers "all conditions of the
# position"). Conservative on purpose; anything missed is user-editable.
_CONDITION_RX = re.compile(
    r"\b(travel|telecommut\w*|remote work|work from home|relocat\w*|"
    r"on[- ]?call)\b", re.I)


def is_condition_of_employment(text: str) -> bool:
    return bool(_CONDITION_RX.search(text or ""))


def months_phrase(months: int | None) -> str | None:
    """72 -> '6 years'; 30 -> '30 months'; None -> None."""
    if not months or months <= 0:
        return None
    if months % 12 == 0:
        years = months // 12
        return f"{years} year{'s' if years != 1 else ''}"
    return f"{months} months"


def skill_question(item: str, months: int | None = None) -> str:
    """Wrap one atomic special requirement in the template question form.

    When the PWD binds a duration to the item (required_months from the
    extractor), the question carries it: "Does the applicant's experience
    involve 2 years of X?" — unless the item text already leads with its
    own duration (e.g. "4 years analyzing data...")."""
    text = " ".join((item or "").split()).rstrip(".;,")
    if not text:
        return ""
    # Lowercase only sentence-case gerund phrases ("Translating analysis…");
    # leave proper nouns and product names alone (Snowflake, Spring Boot, P&L).
    words = text.split(" ", 2)
    if (len(words) > 1 and words[0][:1].isupper()
            and words[0].lower().endswith("ing") and words[0][1:].islower()
            and words[1][:1].islower()):
        text = text[0].lower() + text[1:]
    already_dated = bool(re.match(r"^\d+\s*(year|month)", text, re.I))
    if months and not already_dated:
        dur = months_phrase(months)
        # "experience in X" -> "2 years in X" (avoid "...involve 2 years of experience in...")
        text = re.sub(r"^experience\s+(?=(in|with|as)\b)", "", text, flags=re.I)
        joiner = "of " if not text.startswith(("in ", "with ", "as ")) else ""
        text = f"{dur} {joiner}{text}"
    return f"Does the applicant's experience involve {text}?"


def default_threshold(n_skills: int) -> int:
    """At least half of the special skills, rounded down; minimum 1."""
    return max(1, n_skills // 2) if n_skills > 0 else 0


def config_from_pwd(pwd: dict[str, Any]) -> dict[str, Any]:
    """Map extract_pwd_requirements() output to an editable spreadsheet config."""
    req = pwd.get("requirements") or {}
    primary = req.get("primary") or {}
    alt = req.get("alternative")

    def _edu_q(route: dict[str, Any], label: str) -> str:
        level = (route.get("education_level") or "a").strip()
        # The template supplies its own "or a related field of study" tail;
        # drop any equivalent entry the extractor carried in from the form.
        majors = [m for m in (route.get("fields_of_study") or [])
                  if not re.match(r"^(or\s+)?(a\s+)?related\s+field(s)?"
                                  r"(\s+of\s+study)?$", m.strip(), re.I)]
        art = "an" if level[:1].lower() in "aeiou" else "a"
        field = ", ".join(majors) if majors else "the required field"
        return (f"{label} EDUCATION REQUIREMENT: Does the applicant have {art} "
                f"{level} degree in {field}, or a related field of study?")

    def _exp_q(route: dict[str, Any], label: str) -> str:
        dur = months_phrase(route.get("experience_months")) or "the required experience"
        occ = (route.get("experience_occupation") or "the offered position").strip()
        # Same de-duplication for the "or related position/occupation" tail.
        occ = re.sub(r"[,.]?\s*(or\s+)?(a\s+)?related\s+position(s)?"
                     r"(\s*/\s*occupation(s)?)?[.\s]*$", "", occ, flags=re.I)
        occ = occ.rstrip(".,; ")
        return (f"{label} EXPERIENCE REQUIREMENT: Does the applicant have {dur} "
                f"of experience as a {occ}, or related position/occupation?")

    def _skill_q(item: Any) -> str:
        text = (item.get("text") if isinstance(item, dict) else str(item)) or ""
        if is_condition_of_employment(text):
            return ""
        months = item.get("required_months") if isinstance(item, dict) else None
        return skill_question(text, months=months)

    conditions: list[str] = []
    for item in (primary.get("special_requirements") or []):
        text = (item.get("text") if isinstance(item, dict) else str(item)) or ""
        if is_condition_of_employment(text):
            conditions.append(text)

    skills: list[str] = []
    for item in (primary.get("special_requirements") or []):
        q = _skill_q(item)
        if q and q not in skills:
            skills.append(q)
    if alt and (alt.get("special_requirements_mode") == "replace"):
        for item in (alt.get("special_requirements") or []):
            q = _skill_q(item)
            if q and q not in skills:
                skills.append(q)

    config: dict[str, Any] = {
        "job_title": pwd.get("job_title") or "",
        "req_number": "",
        "primary": {
            "education_question": _edu_q(primary, "PRIMARY"),
            "experience_question": _exp_q(primary, "PRIMARY"),
        },
        "alternative": None,
        "special_skills": skills[:MAX_SKILLS],
        "conditions_excluded": conditions,
        "highlight_rule": {
            "enabled": True,
            "threshold": default_threshold(min(len(skills), MAX_SKILLS)),
        },
    }
    if alt:
        config["alternative"] = {
            "education_question": _edu_q(alt, "ALTERNATIVE"),
            "experience_question": _exp_q(alt, "ALTERNATIVE"),
        }
    return config


def _instructions(has_alt: bool, rule_on: bool, n_skills: int, threshold: int) -> str:
    steps = ["1. Paste the applicant information from the Workday Output Report "
             "into this spreadsheet."]
    n = 2
    steps.append(f"{n}. Evaluate the Primary Requirements."); n += 1
    if has_alt:
        steps.append(f"{n}. Evaluate the Alternative Requirements."); n += 1
    if n_skills:
        steps.append(f"{n}. Evaluate the Special Skills Requirements."); n += 1
    routes = ("either the Primary or Alternative Requirements" if has_alt
              else "the Primary Requirements")
    if rule_on and n_skills:
        steps.append(
            f"{n}. If the applicant meets {routes} and at least {threshold} of the "
            f"{n_skills} Special Skills Requirements, the spreadsheet will recommend "
            "sending the Pre-Screen Questionnaire (the applicant row will be "
            "highlighted).")
        n += 1
    elif rule_on:
        steps.append(f"{n}. If the applicant meets {routes}, the spreadsheet will "
                     "recommend sending the Pre-Screen Questionnaire (the applicant "
                     "row will be highlighted).")
        n += 1
    else:
        steps.append(f"{n}. If the applicant meets {routes} and the Special Skills "
                     "Requirements, mark the Pre-Screen Questionnaire column "
                     f"\"{SEND}\".")
        n += 1
    steps.append(f"{n}. If recommended, send out the Pre-Screen Questionnaire to the "
                 "applicant and complete the related outreach fields."); n += 1
    steps.append(f"{n}. Complete the remaining evaluation sections after receiving "
                 "the applicant's response (if applicable)."); n += 1
    steps.append(f"{n}. Complete the evaluator's conclusion."); n += 1
    steps.append(f"{n}. If the evaluator concludes that the applicant meets all "
                 "requirements, use the Manager Screening Matrix (separate document) "
                 "to contact and interview the applicant.")
    return "Evaluation Process:\n" + "\n".join(steps)


def _banner(ws, c1: int, c2: int, text: str, hexfill: str):
    cell = ws.cell(row=1, column=c1, value=text)
    cell.font = Font(name="Calibri", size=11, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=hexfill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if c2 > c1:
        ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)


def _header(ws, col: int, text: str, width: float = 22.0):
    cell = ws.cell(row=2, column=col, value=text)
    cell.font = Font(name="Calibri", size=9, bold=True)
    cell.fill = PatternFill("solid", fgColor=FILL_GREY_LIGHT)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = BORDER
    ws.column_dimensions[get_column_letter(col)].width = width


def build_workbook(config: dict[str, Any], rows: int = DEFAULT_ROWS) -> Workbook:
    """Build the applicant evaluation workbook from a validated config."""
    rows = max(10, min(int(rows or DEFAULT_ROWS), 5000))
    primary = config.get("primary") or {}
    alt = config.get("alternative")
    skills = [s for s in (config.get("special_skills") or []) if str(s).strip()]
    skills = skills[:MAX_SKILLS]
    rule = config.get("highlight_rule") or {}
    rule_on = bool(rule.get("enabled"))
    threshold = int(rule.get("threshold") or default_threshold(len(skills)))
    threshold = max(0, min(threshold, len(skills)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Applicant Evaluation Form"

    # ── Column layout (dynamic) ─────────────────────────────────────────────
    col = 1
    col_paste = col; col += 1          # A
    col_instr = col; col += 1          # B
    col_p1, col_p2 = col, col + 1; col += 2
    if alt:
        col_a1, col_a2 = col, col + 1; col += 2
    else:
        col_a1 = col_a2 = None
    skill_cols = list(range(col, col + len(skills))); col += len(skills)
    col_q = col; col += 1
    col_contact, col_sent, col_follow, col_resp = col, col + 1, col + 2, col + 3
    col += 4
    col_vresp = col; col += 1
    col_w, col_x, col_y, col_z, col_aa = col, col + 1, col + 2, col + 3, col + 4
    col += 5
    last_col = col - 1
    last_letter = get_column_letter(last_col)
    q_letter = get_column_letter(col_q)

    # ── Row 1: paste zone, instructions, section banners ────────────────────
    a1 = ws.cell(row=1, column=col_paste, value=(
        "[Paste copied columns with applicant information from Workday Output "
        "Report here.]"))
    a1.font = Font(name="Calibri", size=11, bold=True, color="FF0000")
    a1.fill = PatternFill("solid", fgColor=FILL_GREY_LIGHT)
    a1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=1, start_column=col_paste, end_row=2, end_column=col_paste)
    ws.column_dimensions[get_column_letter(col_paste)].width = 39

    b1 = ws.cell(row=1, column=col_instr, value=_instructions(
        bool(alt), rule_on, len(skills), threshold))
    b1.font = Font(name="Calibri", size=10, bold=True)
    b1.fill = PatternFill("solid", fgColor=FILL_GREY_LIGHT)
    b1.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=1, start_column=col_instr, end_row=2, end_column=col_instr)
    ws.column_dimensions[get_column_letter(col_instr)].width = 84

    _banner(ws, col_p1, col_p2,
            "PRIMARY REQUIREMENTS: (Determine whether the applicant meets the "
            "Primary Education and Experience Requirements.)", FILL_PRIMARY)
    if alt:
        _banner(ws, col_a1, col_a2,
                "ALTERNATIVE REQUIREMENTS: (Determine whether the applicant meets "
                "the Alternative Education and Experience Requirements.)", FILL_ALT)
    if skills:
        pct = (f"at least {threshold} of the {len(skills)} Special Skills "
               "Requirements") if rule_on else "the Special Skills Requirements"
        routes = ("either the Primary or Alternative Requirements" if alt
                  else "the Primary Requirements")
        _banner(ws, skill_cols[0], skill_cols[-1],
                "SPECIAL SKILLS REQUIREMENTS: (Determine whether the applicant "
                f"meets the Special Skills Requirements. If the applicant meets "
                f"{pct} and {routes}, send the applicant the pre-screen "
                "questionnaire.)", FILL_SKILLS)
    _banner(ws, col_q, col_resp,
            "Pre-Screen Questionnaire Outreach: (Complete only if a pre-screen "
            "questionnaire was sent to the applicant.)", FILL_OUTREACH)
    _banner(ws, col_vresp, col_vresp, "Pre-Screen Questionnaire Responses:",
            FILL_RESPONSES)
    _banner(ws, col_w, col_aa,
            "Evaluator's Conclusion: (Complete after reviewing all evaluation "
            "criteria and applicant responses.)", FILL_CONCLUDE)

    # ── Row 2: per-column question headers ──────────────────────────────────
    _header(ws, col_p1, primary.get("education_question") or
            "PRIMARY EDUCATION REQUIREMENT:", 25)
    _header(ws, col_p2, primary.get("experience_question") or
            "PRIMARY EXPERIENCE REQUIREMENT:", 25)
    if alt:
        _header(ws, col_a1, alt.get("education_question") or
                "ALTERNATIVE EDUCATION REQUIREMENT:", 25)
        _header(ws, col_a2, alt.get("experience_question") or
                "ALTERNATIVE EXPERIENCE REQUIREMENT:", 25)
    for i, c in enumerate(skill_cols):
        _header(ws, c, skills[i], 21)

    q_header = "Pre-Screen Questionnaire Required?:"
    if rule_on:
        q_header += "\n\n" + DISCLAIMER
    _header(ws, col_q, q_header, 33)
    _header(ws, col_contact, "Contact Method (e.g., email, phone, etc.):", 25)
    _header(ws, col_sent, "Pre-Screen Questionnaire Sent Date:", 22)
    _header(ws, col_follow, "Follow Up Date(s):", 22)
    _header(ws, col_resp, "Applicant Response Received?:", 22)

    routes_txt = ("either the Primary or Alternative Requirements" if alt
                  else "the Primary Requirements")
    _header(ws, col_vresp,
            "Based on the applicant's Pre-Screen Questionnaire responses, does "
            f"the applicant meet all the position requirements including "
            f"{routes_txt}, all Special Skills Requirements, and all conditions "
            "of the position?", 29)
    _header(ws, col_w,
            "Can training be provided within a reasonable amount of time for any "
            "experiential or skill requirement which the Applicant has been found "
            "deficient? WARNING: If yes, please explain and call our staff to "
            "discuss.", 28)
    _header(ws, col_x,
            "Does the Evaluator believe that an INTERVIEW is recommended and/or "
            "necessary to accurately assess the Applicant's competence levels in "
            "the specified skill areas and/or to assess whether the Applicant "
            "satisfies the basic academic and/or experience requirements?", 28)
    _header(ws, col_y,
            "Is the Evaluator concluding that the Applicant is not qualified for "
            "the subject job opening based solely upon valid job-related reasons "
            "derived from the criteria in the approved job offer?", 28)
    _header(ws, col_z,
            "Disposition (e.g., Qualified (Proceed to Interview), Not Qualified "
            "(Missing Education), Not Qualified (Missing Experience), Not "
            "Qualified (Missing Skill), Unresponsive, Requires Immigration "
            "Sponsorship, etc.):", 28)
    _header(ws, col_aa, "Any additional comments regarding the applicant:", 24)

    ws.row_dimensions[1].height = 62
    ws.row_dimensions[2].height = 150
    ws.freeze_panes = "C3"

    # ── Data rows ───────────────────────────────────────────────────────────
    first_row, last_row = 3, 2 + rows
    yn_first = get_column_letter(col_p1)
    yn_last = get_column_letter(skill_cols[-1] if skill_cols else
                                (col_a2 if alt else col_p2))
    dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
    dv.error = "Choose YES or NO."
    ws.add_data_validation(dv)
    dv.add(f"{yn_first}{first_row}:{yn_last}{last_row}")

    if skill_cols and rule_on:
        s_first = get_column_letter(skill_cols[0])
        s_last = get_column_letter(skill_cols[-1])
        p1, p2 = get_column_letter(col_p1), get_column_letter(col_p2)
        for r in range(first_row, last_row + 1):
            routes_f = f'AND({p1}{r}="YES",{p2}{r}="YES")'
            if alt:
                a1l, a2l = get_column_letter(col_a1), get_column_letter(col_a2)
                routes_f = (f'OR({routes_f},AND({a1l}{r}="YES",{a2l}{r}="YES"))')
            skills_f = f'COUNTIF({s_first}{r}:{s_last}{r},"YES")>={threshold}'
            ws.cell(row=r, column=col_q, value=(
                f'=IF(AND({routes_f},{skills_f}),"{SEND}","{NOSEND}")'))
    elif rule_on:
        p1, p2 = get_column_letter(col_p1), get_column_letter(col_p2)
        for r in range(first_row, last_row + 1):
            routes_f = f'AND({p1}{r}="YES",{p2}{r}="YES")'
            if alt:
                a1l, a2l = get_column_letter(col_a1), get_column_letter(col_a2)
                routes_f = (f'OR({routes_f},AND({a1l}{r}="YES",{a2l}{r}="YES"))')
            ws.cell(row=r, column=col_q,
                    value=f'=IF({routes_f},"{SEND}","{NOSEND}")')
    else:
        dv_q = DataValidation(type="list", formula1=f'"{SEND},{NOSEND}"',
                              allow_blank=True)
        ws.add_data_validation(dv_q)
        dv_q.add(f"{q_letter}{first_row}:{q_letter}{last_row}")

    bold = Font(name="Calibri", size=11, bold=True)
    for r in range(first_row, last_row + 1):
        ws.cell(row=r, column=col_q).font = bold

    # ── Highlight rule ──────────────────────────────────────────────────────
    if rule_on:
        ws.conditional_formatting.add(
            f"A{first_row}:{last_letter}{last_row}",
            FormulaRule(formula=[f'${q_letter}{first_row}="{SEND}"'],
                        fill=PatternFill("solid", bgColor=FILL_HIGHLIGHT),
                        stopIfTrue=False))
    return wb


def workbook_bytes(config: dict[str, Any], rows: int = DEFAULT_ROWS) -> bytes:
    wb = build_workbook(config, rows=rows)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def suggested_filename(config: dict[str, Any]) -> str:
    title = _sanitize(config.get("job_title") or "Position")
    req = _sanitize(config.get("req_number") or "")
    parts = ["Applicant Evaluation Spreadsheet", title]
    if req:
        parts.append(f"Req {req}")
    return " - ".join(parts).replace(" ", "_") + ".xlsx"
