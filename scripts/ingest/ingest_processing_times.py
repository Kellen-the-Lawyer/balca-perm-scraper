#!/usr/bin/env python3
"""Load dedicated USCIS processing-time and I-129 classification observations.

The generic USCIS statistics importer intentionally handles heterogeneous files.
These report families need exact parsers because their multi-row headers otherwise
shift dimensions and mislabel USCIS months as days.

Usage:
    DATABASE_URL=postgresql://perm@127.0.0.1:5433/perm_decisions \
      venv/bin/python3 scripts/ingest/ingest_processing_times.py
"""

from __future__ import annotations

import calendar
import csv
import hashlib
import json
import os
import re
from datetime import date
from pathlib import Path

import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
load_dotenv(ROOT / ".env")
DB_URL = os.environ.get("DATABASE_URL", "postgresql://perm@127.0.0.1:5433/perm_decisions")
SCHEMA = ROOT / "schema" / "processing_times_schema.sql"
I140_QUARTERLY = ROOT / "data" / "processed" / "uscis_i140_quarterly.json"

MONTHS = {name.lower(): number for number, name in enumerate(calendar.month_name) if name}
MONTH_REPORT_RE = re.compile(r"Application Processing Data for\s+([A-Za-z]+),?\s+(\d{4})", re.I)


def clean_number(value):
    if value is None:
        return None
    text = re.sub(r"[,$%\s]", "", str(value))
    if not text or text.lower() in {"n/a", "na", "-"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def stable_id(*parts):
    raw = "|".join("" if part is None else str(part) for part in parts)
    return hashlib.sha256(raw.encode()).hexdigest()[:40]


def month_bounds(year: int, month: int):
    return date(year, month, 1), date(year, month, calendar.monthrange(year, month)[1])


def fiscal_month_date(fiscal_year: int, month_name: str):
    month = MONTHS[month_name.strip().lower()]
    year = fiscal_year - 1 if month >= 10 else fiscal_year
    return month_bounds(year, month)


def fiscal_quarter_bounds(fiscal_year: int, quarter: int):
    start_year, start_month = {
        1: (fiscal_year - 1, 10), 2: (fiscal_year, 1),
        3: (fiscal_year, 4), 4: (fiscal_year, 7),
    }[quarter]
    end_year, end_month = {
        1: (fiscal_year - 1, 12), 2: (fiscal_year, 3),
        3: (fiscal_year, 6), 4: (fiscal_year, 9),
    }[quarter]
    return date(start_year, start_month, 1), date(
        end_year, end_month, calendar.monthrange(end_year, end_month)[1]
    )


def observation(**kwargs):
    identifying = (
        kwargs["agency"], kwargs["series_key"], kwargs["period_start"],
        kwargs["period_end"], kwargs["metric_name"], kwargs["statistic"],
    )
    kwargs.setdefault("stable_id", stable_id(*identifying))
    kwargs.setdefault("metadata", {})
    return kwargs


def parse_monthly_processing_csv(path: Path, report_id: int, title: str, source_url: str):
    match = MONTH_REPORT_RE.search(title)
    if not match:
        return []
    month = MONTHS[match.group(1).lower()]
    year = int(match.group(2))
    start, end = month_bounds(year, month)

    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        rows = list(csv.reader(handle))
    header_idx = next((i for i, row in enumerate(rows) if row and row[0].strip() == "Form Number"), None)
    if header_idx is None:
        return []

    out = []
    for row in rows[header_idx + 1:]:
        if len(row) < 8:
            continue
        form = row[0].strip()
        description = row[1].strip()
        value = clean_number(row[7])
        if not re.fullmatch(r"[A-Z]-?\d+[A-Z]?", form) or value is None:
            continue
        classification = description or "All classifications"
        key = f"uscis-monthly:{form}:{classification.lower()}"
        out.append(observation(
            agency="USCIS", series_key=key,
            series_label=f"{form} — {classification}", form_type=form,
            classification=classification, office="All USCIS offices",
            period_start=start, period_end=end, period_granularity="month",
            metric_name="processing_time", statistic="average", value=value,
            unit="months", case_count=None, source_name="USCIS monthly appropriations report",
            source_url=source_url, source_report_id=report_id, source_file=str(path),
            metadata={"methodology": "Average receipt-to-completion time for cases completed in the month."},
        ))
    return out


def _classification_from_sheet(sheet_name: str):
    suffix = sheet_name.split("_")[-1].upper()
    mapping = {
        "H1B": "H-1B", "H2A": "H-2A", "H2B": "H-2B", "BLANKET L": "Blanket L",
        "BLANKET_L": "Blanket L", "L1A": "L-1A", "L1B": "L-1B", "L1": "L-1",
        "O": "O", "P": "P", "R1": "R-1", "TN": "TN",
    }
    normalized = sheet_name.upper()
    for token, label in mapping.items():
        if normalized.endswith(token) or f"_{token}_" in normalized:
            return label
    return suffix.replace("FY24Q4", "").strip("_")


def parse_i129_workbook(path: Path, report_id: int, source_url: str):
    workbook = load_workbook(path, read_only=True, data_only=True)
    out = []
    for sheet in workbook.worksheets:
        if "note" in sheet.title.lower():
            continue
        classification = _classification_from_sheet(sheet.title)
        fiscal_year = None
        for values in sheet.iter_rows(min_row=6, values_only=True):
            first, month_name = values[0], values[1]
            if isinstance(first, (int, float)):
                fiscal_year = int(first)
            if not fiscal_year or not isinstance(month_name, str) or month_name.strip().lower() not in MONTHS:
                continue
            start, end = fiscal_month_date(fiscal_year, month_name)
            received = clean_number(values[2])
            completed = clean_number(values[5])
            # Newer files insert Pending before approval rate; identify rates by magnitude.
            candidates = [clean_number(v) for v in values[6:12]]
            rates = [v for v in candidates if v is not None and 0 <= v <= 1]
            approval_rate = rates[0] if rates else None
            rfe_rate = rates[1] if len(rates) > 1 else None
            metrics = (
                ("received", "count", received, "cases"),
                ("completed", "count", completed, "cases"),
                ("approval_rate", "rate", approval_rate, "proportion"),
                ("rfe_rate", "rate", rfe_rate, "proportion"),
            )
            for metric_name, statistic, value, unit in metrics:
                if value is None:
                    continue
                key = f"uscis-i129-context:{classification.lower()}"
                out.append(observation(
                    agency="USCIS", series_key=key,
                    series_label=f"I-129 {classification}", form_type="I-129",
                    classification=classification, office="All USCIS offices",
                    period_start=start, period_end=end, period_granularity="month",
                    metric_name=metric_name, statistic=statistic, value=value, unit=unit,
                    case_count=int(completed) if completed is not None else None,
                    source_name="USCIS I-129 case status and RFE report",
                    source_url=source_url, source_report_id=report_id, source_file=str(path),
                    metadata={"sheet": sheet.title},
                ))
    return out


def parse_all_forms_workbook(path: Path, report_id: int, fiscal_year: int, quarter: int, source_url: str):
    """Parse USCIS's quarterly median processing-time column (months)."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    start_month = {1: (fiscal_year - 1, 10), 2: (fiscal_year, 1), 3: (fiscal_year, 4), 4: (fiscal_year, 7)}[quarter]
    end_month = {1: (fiscal_year - 1, 12), 2: (fiscal_year, 3), 3: (fiscal_year, 6), 4: (fiscal_year, 9)}[quarter]
    start = date(start_month[0], start_month[1], 1)
    end = date(end_month[0], end_month[1], calendar.monthrange(*end_month)[1])
    out = []
    for values in sheet.iter_rows(min_row=6, values_only=True):
        form = str(values[0] or "").strip()
        description = str(values[1] or "").strip()
        value = clean_number(values[7] if len(values) > 7 else None)
        match = re.match(r"([A-Z]-?\d+[A-Z]?)", form)
        if not match or value is None:
            continue
        form_type = match.group(1)
        classification = description or "All classifications"
        key = f"uscis-quarterly:{form_type}:{classification.lower()}"
        out.append(observation(
            agency="USCIS", series_key=key,
            series_label=f"{form_type} — {classification}", form_type=form_type,
            classification=classification, office="All USCIS offices",
            period_start=start, period_end=end, period_granularity="quarter",
            metric_name="processing_time", statistic="median", value=value,
            unit="months", case_count=None, source_name="USCIS All Forms quarterly report",
            source_url=source_url, source_report_id=report_id, source_file=str(path),
            metadata={"fiscal_year": fiscal_year, "fiscal_quarter": quarter,
                      "methodology": "Median receipt-to-completion time for cases completed in the quarter."},
        ))
    return out


OFFICIAL_I129_SNAPSHOTS = [
    ("137-H1B1", "H-1B — visa issued abroad", 9.5),
    ("137-H1B2", "H-1B — change of status", 9.5),
    ("137-H1B3", "H-1B — extension of stay", 10.5),
    ("137-O", "O — extraordinary ability", 13.0),
]


def official_snapshot_rows():
    observed = date(2026, 7, 16)
    return [observation(
        agency="USCIS", series_key=f"uscis-current:I-129:{code}", series_label=label,
        form_type="I-129", classification=code, office="Service Center Operations (SCOPS)",
        period_start=observed, period_end=observed, period_granularity="snapshot",
        metric_name="processing_time", statistic="p80", value=value, unit="months",
        case_count=None, source_name="USCIS Case Processing Times",
        source_url="https://egov.uscis.gov/processing-times/", source_report_id=None,
        source_file=None, metadata={
            "observed_on": observed.isoformat(),
            "methodology": "Time in which USCIS completed 80% of adjudicated cases over the prior six months; premium cases excluded.",
        },
    ) for code, label, value in OFFICIAL_I129_SNAPSHOTS]


def i140_quarterly_rows():
    """Load normalized category outcomes extracted from USCIS quarterly workbooks."""
    if not I140_QUARTERLY.exists():
        return []
    out = []
    for record in json.loads(I140_QUARTERLY.read_text()):
        fiscal_year = int(record["fiscal_year"])
        quarter = int(record["fiscal_quarter"])
        classification = record["classification"]
        start, end = fiscal_quarter_bounds(fiscal_year, quarter)
        decided = float(record["approved"]) + float(record["denied"])
        metrics = [
            ("received", "count", record["received"], "cases"),
            ("approved", "count", record["approved"], "cases"),
            ("denied", "count", record["denied"], "cases"),
            ("pending", "snapshot_count", record["pending"], "cases"),
        ]
        if decided:
            metrics.extend([
                ("decision_approval_rate", "rate", float(record["approved"]) / decided, "proportion"),
                ("decision_denial_rate", "rate", float(record["denied"]) / decided, "proportion"),
            ])
        for metric_name, statistic, value, unit in metrics:
            out.append(observation(
                agency="USCIS", series_key=f"uscis-i140-context:{classification.lower()}",
                series_label=f"I-140 {classification}", form_type="I-140",
                classification=classification, office="All USCIS offices",
                period_start=start, period_end=end, period_granularity="quarter",
                metric_name=metric_name, statistic=statistic, value=value, unit=unit,
                case_count=int(decided),
                source_name="USCIS Form I-140 by Fiscal Year, Quarter and Case Status",
                source_url="https://www.uscis.gov/tools/reports-and-studies/immigration-and-citizenship-data",
                source_report_id=None, source_file=record["source_file"],
                metadata={
                    "fiscal_year": fiscal_year, "fiscal_quarter": quarter,
                    "source_label": record["source_label"],
                    "decision_rate_methodology": "Approved or denied cases completed in the quarter; not a filing cohort approval rate.",
                },
            ))
    return out


INSERT_SQL = """
INSERT INTO processing_time_observations (
    stable_id, agency, series_key, series_label, program, form_type, classification,
    office, period_start, period_end, period_granularity, metric_name, statistic,
    value, lower_value, upper_value, unit, case_count, source_name, source_url,
    source_report_id, source_file, metadata
) VALUES (
    %(stable_id)s, %(agency)s, %(series_key)s, %(series_label)s, %(program)s,
    %(form_type)s, %(classification)s, %(office)s, %(period_start)s, %(period_end)s,
    %(period_granularity)s, %(metric_name)s, %(statistic)s, %(value)s,
    %(lower_value)s, %(upper_value)s, %(unit)s, %(case_count)s, %(source_name)s,
    %(source_url)s, %(source_report_id)s, %(source_file)s, %(metadata)s
)
ON CONFLICT (stable_id) DO UPDATE SET
    value=EXCLUDED.value, lower_value=EXCLUDED.lower_value,
    upper_value=EXCLUDED.upper_value, case_count=EXCLUDED.case_count,
    metadata=EXCLUDED.metadata, ingested_at=NOW()
"""


def load():
    conn = psycopg2.connect(DB_URL)
    observations = []
    with conn, conn.cursor() as cur:
        cur.execute(SCHEMA.read_text())
        cur.execute("""
            SELECT c.id, c.title, c.file_url, c.fiscal_year, c.quarter, f.local_path
            FROM uscis_report_catalog c
            JOIN uscis_report_files f ON f.report_id=c.id
            WHERE f.download_status='done' AND f.local_path IS NOT NULL
              AND (c.title ILIKE 'FY22 Appropriations Reporting Requirement - Application Processing Data%'
                   OR c.title ILIKE 'Nonimmigrant Worker Petitions by Case Status and Request for Evidence%'
                   OR c.title ILIKE 'All USCIS Application and Petition Form Types%')
            ORDER BY c.published_date
        """)
        reports = cur.fetchall()

        # Monthly processing reports are independent observations; load all 42.
        for report_id, title, url, _fy, _quarter, local_path in reports:
            path = Path(local_path)
            if path.suffix.lower() == ".csv" and title.startswith("FY22 Appropriations"):
                observations.extend(parse_monthly_processing_csv(path, report_id, title, url))

        # Use non-overlapping classification workbooks: full FY2017-24, FY2025, FY2026 Q1.
        selected = []
        for report in reports:
            report_id, title, url, fy, quarter, local_path = report
            path = Path(local_path)
            if not title.startswith("Nonimmigrant Worker") or path.suffix.lower() != ".xlsx":
                continue
            if (fy == 2024 and quarter == 4) or (fy == 2025 and quarter == 4) or (fy == 2026 and quarter == 1):
                selected.append(report)
        for report_id, _title, url, _fy, _quarter, local_path in selected:
            observations.extend(parse_i129_workbook(Path(local_path), report_id, url))

        # Structured All Forms workbooks begin in FY2024 and provide quarterly medians.
        for report_id, title, url, fy, quarter, local_path in reports:
            path = Path(local_path)
            if (title.startswith("All USCIS Application") and path.suffix.lower() == ".xlsx"
                    and fy and quarter):
                observations.extend(parse_all_forms_workbook(path, report_id, int(fy), int(quarter), url))

        observations.extend(official_snapshot_rows())
        observations.extend(i140_quarterly_rows())
        for item in observations:
            item.setdefault("program", None)
            item.setdefault("form_type", None)
            item.setdefault("classification", None)
            item.setdefault("office", None)
            item.setdefault("lower_value", None)
            item.setdefault("upper_value", None)
            item.setdefault("case_count", None)
            item.setdefault("source_report_id", None)
            item.setdefault("source_file", None)
            item["metadata"] = psycopg2.extras.Json(item.get("metadata") or {})
        psycopg2.extras.execute_batch(cur, INSERT_SQL, observations, page_size=1000)
    conn.close()
    print(f"Loaded {len(observations):,} processing-time/context observations")


if __name__ == "__main__":
    load()
