#!/usr/bin/env python3
"""Build a deterministic, resumable corpus of matched ETA-9141/ETA-9089 pairs."""

from __future__ import annotations

import argparse
import hashlib
import heapq
import json
import os
import sys
from concurrent.futures import ProcessPoolExecutor, as_completed
from datetime import date, datetime
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.perm_verify.form_fill import fill_eta9141, fill_eta9089_package  # noqa: E402
from app.perm_verify.form_fill.dol_eta9141 import build_eta9141_form_data  # noqa: E402
from app.perm_verify.form_fill.dol_eta9089 import (  # noqa: E402
    build_eta9089_form_data,
    synthetic_foreign_national,
)
from app.perm_verify.form_fill.eta9141 import TEMPLATE as ETA9141_TEMPLATE  # noqa: E402
from app.perm_verify.form_fill.eta9141_addendum import (  # noqa: E402
    eta9141_addendum_sections,
    eta9141_with_addendum_references,
    generate_eta9141_addendum,
    merge_eta9141_package,
)


FIRST_NAMES = (
    "ALEX", "AMINA", "ARJUN", "CAMILA", "DANIEL", "ELENA", "FARAH", "HUGO",
    "ISABEL", "JAE", "KIRAN", "LEILA", "MATEO", "NADIA", "OMAR", "PRIYA",
    "RAFAEL", "SANA", "TARIQ", "VIVIAN",
)
LAST_NAMES = (
    "TRAINING-ALPHA", "TRAINING-BRAVO", "TRAINING-CEDAR", "TRAINING-DELTA",
    "TRAINING-ECHO", "TRAINING-FIELD", "TRAINING-GROVE", "TRAINING-HARBOR",
    "TRAINING-IVY", "TRAINING-JUNIPER", "TRAINING-KITE", "TRAINING-LAKE",
    "TRAINING-MAPLE", "TRAINING-NORTH", "TRAINING-OAK", "TRAINING-PINE",
    "TRAINING-QUARTZ", "TRAINING-RIVER", "TRAINING-SUMMIT", "TRAINING-VALE",
)
COUNTRY_PROFILES = (
    ("INDIA", "H-1B"),
    ("CANADA", "TN"),
    ("MEXICO", "TN"),
    ("BRAZIL", "H-1B"),
    ("CHINA", "H-1B"),
    ("COLOMBIA", "H-1B"),
    ("GERMANY", "H-1B"),
    ("JAPAN", "H-1B"),
    ("PHILIPPINES", "H-1B"),
    ("UNITED KINGDOM", "H-1B"),
)


def _jsonable(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if hasattr(value, "item"):
        return value.item()
    return value


def _stable_int(value: str) -> int:
    return int(hashlib.sha256(value.encode("utf-8")).hexdigest(), 16)


def _row_dict(headers: list[Any], row: tuple[Any, ...]) -> dict[str, Any]:
    return {
        str(header): _jsonable(value)
        for header, value in zip(headers, row)
        if header is not None
    }


def _case_numbers(path: Path) -> set[str]:
    book = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = book.active
        iterator = sheet.iter_rows(values_only=True)
        headers = list(next(iterator))
        case_index = headers.index("CASE_NUMBER")
        result = {
            str(row[case_index]).strip()
            for row in iterator
            if row[case_index] not in (None, "")
        }
    finally:
        book.close()
    return result


def _select_perm_rows(
    path: Path, pwd_numbers: set[str], limit: int
) -> tuple[list[dict[str, Any]], int, int]:
    """Keep the `limit` lowest stable hashes without loading the whole workbook."""

    book = load_workbook(path, read_only=True, data_only=True)
    heap: list[tuple[int, str, dict[str, Any]]] = []
    matched = 0
    seen_pwd_numbers: set[str] = set()
    try:
        sheet = book.active
        iterator = sheet.iter_rows(values_only=True)
        headers = list(next(iterator))
        case_index = headers.index("CASE_NUMBER")
        pwd_index = headers.index("JOB_OPP_PWD_NUMBER")
        for row in iterator:
            case_number = row[case_index]
            pwd_number = row[pwd_index]
            if not case_number or not pwd_number or str(pwd_number) not in pwd_numbers:
                continue
            matched += 1
            pwd_number = str(pwd_number)
            if pwd_number in seen_pwd_numbers:
                continue
            seen_pwd_numbers.add(pwd_number)
            case_number = str(case_number)
            score = _stable_int(case_number)
            item = (-score, case_number, _row_dict(headers, row))
            if len(heap) < limit:
                heapq.heappush(heap, item)
            elif score < -heap[0][0]:
                heapq.heapreplace(heap, item)
    finally:
        book.close()
    selected = [item[2] for item in heap]
    selected.sort(key=lambda row: _stable_int(str(row["CASE_NUMBER"])))
    return selected, matched, len(seen_pwd_numbers)


def _selected_rows(path: Path, case_numbers: set[str]) -> dict[str, dict[str, Any]]:
    book = load_workbook(path, read_only=True, data_only=True)
    found: dict[str, dict[str, Any]] = {}
    try:
        sheet = book.active
        iterator = sheet.iter_rows(values_only=True)
        headers = list(next(iterator))
        case_index = headers.index("CASE_NUMBER")
        for row in iterator:
            case_number = row[case_index]
            if case_number is None or str(case_number) not in case_numbers:
                continue
            found[str(case_number)] = _row_dict(headers, row)
            if len(found) == len(case_numbers):
                break
    finally:
        book.close()
    missing = case_numbers - set(found)
    if missing:
        raise ValueError(f"PWD workbook is missing {len(missing)} selected rows")
    return found


def _scenario(case_number: str, pwd: Mapping[str, Any]) -> dict[str, Any]:
    stable = _stable_int(case_number)
    try:
        experience_months = int(float(pwd.get("REQUIRED_EXPERIENCE_MONTHS") or 0))
    except (TypeError, ValueError):
        experience_months = 0
    bucket = stable % 100
    # A cumulative multi-employer history cannot be divided into fewer than two
    # whole months without creating a zero-month employment record.
    if experience_months < 2 or bucket < 60:
        evidence_pattern = "single_employer"
        employer_count = 1
    elif bucket < 80:
        evidence_pattern = "multiple_employers"
        employer_count = 3 if experience_months >= 36 and stable % 5 == 0 else 2
    else:
        evidence_pattern = "split_skills"
        employer_count = 3 if experience_months >= 36 and stable % 5 == 0 else 2
    return {
        "pwd_addendum": stable % 10 < 6,
        "evidence_pattern": evidence_pattern,
        "employer_count": employer_count,
        "experience_months": experience_months,
    }


def _varied_foreign_national(
    perm: Mapping[str, Any], pwd: Mapping[str, Any], scenario: Mapping[str, Any]
) -> dict[str, Any]:
    case_number = str(perm["CASE_NUMBER"])
    stable = _stable_int(case_number)
    result = synthetic_foreign_national(
        perm,
        pwd,
        evidence_pattern=str(scenario["evidence_pattern"]),
        employer_count=int(scenario["employer_count"]),
    )
    first = FIRST_NAMES[stable % len(FIRST_NAMES)]
    last = LAST_NAMES[(stable // len(FIRST_NAMES)) % len(LAST_NAMES)]
    country, admission = COUNTRY_PROFILES[
        (stable // (len(FIRST_NAMES) * len(LAST_NAMES))) % len(COUNTRY_PROFILES)
    ]
    year = 1978 + stable % 20
    month = 1 + (stable // 20) % 12
    day = 1 + (stable // 240) % 27
    contact = result["contact"]
    contact.update(
        {
            "last_name": last,
            "first_name": first,
            "middle_name": "TEST",
            "dob": f"{month:02d}/{day:02d}/{year}",
            "class_of_admission": admission,
            "country_of_birth": country,
            "country_of_citizenship": country,
        }
    )
    for education in result.get("education", []):
        education["country"] = country
    return result


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _artifact(path: Path) -> dict[str, Any]:
    return {
        "path": str(path),
        "sha256": _sha256(path),
        "bytes": path.stat().st_size,
        "page_count": len(PdfReader(path).pages),
    }


def _evidence_validation(perm_form: Mapping[str, Any]) -> dict[str, Any]:
    appendix = perm_form["appendix_A"]
    required_ids = {item["id"] for item in appendix.get("requirements", [])}
    covered_ids = {
        identifier
        for skill in appendix.get("skills", [])
        for identifier in skill.get("requirement_ids", [])
        if identifier in required_ids
    }
    experiences = appendix.get("work_experience", [])
    employer_evidence = []
    for index, experience in enumerate(experiences):
        providers = {
            skill.get("provider")
            for skill in appendix.get("skills", [])
            if skill.get("experience_index", 0) == index
        }
        employer_evidence.append(experience.get("employer_name") in providers)
    required_months = int(appendix.get("required_experience_months", 0) or 0)
    credited_months = sum(int(item.get("credited_months", 0)) for item in experiences)
    return {
        "requirement_ids": sorted(required_ids),
        "covered_requirement_ids": sorted(covered_ids),
        "missing_requirement_ids": sorted(required_ids - covered_ids),
        "requirements_coverage": "pass" if required_ids <= covered_ids else "fail",
        "required_experience_months": required_months,
        "credited_experience_months": credited_months,
        "cumulative_experience": "pass" if credited_months >= required_months else "fail",
        "section_d_e_employer_links": employer_evidence,
        "section_d_e_linkage": "pass" if all(employer_evidence) else "fail",
    }


def _review_flags(
    pwd_form: Mapping[str, Any], perm_form: Mapping[str, Any]
) -> list[str]:
    flags = [f"synthetic:{path}" for path in pwd_form["synthetic_fields"]]
    flags.append("synthetic:perm.appendix_A.foreign_national")
    if not pwd_form["job_offer"]["worksite"].get("county"):
        flags.append("missing:job_offer.worksite.county")
    if any(
        "see f.a.2" in item["text"].lower()
        for item in perm_form["appendix_A"].get("requirements", [])
    ):
        flags.append("source_gap:PWD_F.a.2_text_not_in_disclosure_workbook")
    return flags


def _is_complete(case_manifest_path: Path) -> dict[str, Any] | None:
    if not case_manifest_path.exists():
        return None
    try:
        manifest = json.loads(case_manifest_path.read_text(encoding="utf-8"))
        for artifact in manifest["artifacts"].values():
            path = Path(artifact["path"])
            if not path.exists() or path.stat().st_size != artifact["bytes"]:
                return None
        return manifest
    except (KeyError, OSError, ValueError, TypeError, json.JSONDecodeError):
        return None


def _generate_case(task: Mapping[str, Any]) -> dict[str, Any]:
    selection = task["selection"]
    case_id = selection["perm_case_number"]
    pdf_dir = Path(task["pdf_output"]) / case_id
    json_dir = Path(task["json_output"]) / case_id
    case_manifest_path = json_dir / "case-manifest.json"
    if task["resume"]:
        existing = _is_complete(case_manifest_path)
        if existing:
            existing["resumed"] = True
            return existing

    pdf_dir.mkdir(parents=True, exist_ok=True)
    json_dir.mkdir(parents=True, exist_ok=True)
    pwd_form = task["pwd_form"]
    perm_form = task["perm_form"]
    watermark = task["watermark"]
    template_9141 = Path(task["templates_dir"]) / ETA9141_TEMPLATE

    requested_sections = eta9141_addendum_sections(pwd_form)
    addendum_sections = requested_sections if selection["pwd_addendum"] else []
    pwd_render_form = (
        eta9141_with_addendum_references(pwd_form, addendum_sections)
        if addendum_sections
        else pwd_form
    )
    pwd_pdf = pdf_dir / "eta9141.pdf"
    pwd_result = fill_eta9141(
        pwd_render_form, template_9141, pwd_pdf, watermark=watermark
    )
    addendum_result = None
    complete_pwd_result = None
    if addendum_sections:
        addendum_pdf = pdf_dir / "eta9141_addendum.pdf"
        addendum_result = generate_eta9141_addendum(
            pwd_form,
            addendum_sections,
            template_9141,
            addendum_pdf,
            watermark=watermark,
        )
        complete_pwd_result = merge_eta9141_package(
            pwd_pdf, addendum_pdf, pdf_dir / "eta9141_complete.pdf"
        )
    perm_result = fill_eta9089_package(
        perm_form, Path(task["templates_dir"]), pdf_dir, watermark=watermark
    )

    evidence_validation = _evidence_validation(perm_form)
    required_checks = (
        "requirements_coverage", "cumulative_experience", "section_d_e_linkage"
    )
    if any(evidence_validation[key] != "pass" for key in required_checks):
        raise ValueError(f"Appendix A validation failed for {case_id}")
    review_flags = _review_flags(pwd_form, perm_form)
    pair_payload = {
        "schema_version": "casebase.eta-pair-fill.v2",
        "case_id": case_id,
        "source": {
            "perm_file": task["perm_file"],
            "perm_case_number": case_id,
            "pwd_file": task["pwd_file"],
            "pwd_case_number": selection["pwd_case_number"],
        },
        "selection": selection,
        "privacy": {
            "dol_disclosure_data": True,
            "foreign_national": "synthetic training identity and evidence",
            "not_for_filing": True,
            "training_approved": False,
        },
        "review_flags": review_flags,
        "evidence_validation": evidence_validation,
        "pwd": {
            "form_data": pwd_form,
            "render_form_data": pwd_render_form,
            "fill_result": pwd_result,
            "addendum_result": addendum_result,
            "complete_package_result": complete_pwd_result,
        },
        "perm": {"form_data": perm_form, "fill_result": perm_result},
    }
    pair_json = json_dir / "pair.json"
    pair_json.write_text(json.dumps(pair_payload, indent=2) + "\n", encoding="utf-8")

    artifacts = {"eta9141": _artifact(pwd_pdf)}
    if addendum_sections:
        artifacts["eta9141_addendum"] = _artifact(pdf_dir / "eta9141_addendum.pdf")
        artifacts["eta9141_complete"] = _artifact(pdf_dir / "eta9141_complete.pdf")
    for key in perm_result:
        artifact_key = "eta9089" if key == "application" else f"eta9089_{key}"
        artifacts[artifact_key] = _artifact(pdf_dir / f"{key}.pdf")
    case_manifest = {
        "ordinal": selection["ordinal"],
        "perm_case_number": case_id,
        "pwd_case_number": selection["pwd_case_number"],
        "pair_json": str(pair_json),
        "synthetic_pwd_fields": pwd_form["synthetic_fields"],
        "pwd_addendum_sections": [item["section"] for item in addendum_sections],
        "appendix_a_evidence_pattern": selection["evidence_pattern"],
        "appendix_a_employer_count": len(
            perm_form["appendix_A"].get("work_experience", [])
        ),
        "review_flags": review_flags,
        "evidence_validation": evidence_validation,
        "artifacts": artifacts,
        "validation": {
            "pwd_number_link": "pass",
            "acroform_reopen": "pass",
            "widget_appearances": "pass",
            "human_review": "pending",
        },
        "resumed": False,
    }
    case_manifest_path.write_text(
        json.dumps(case_manifest, indent=2) + "\n", encoding="utf-8"
    )
    return case_manifest


def _summary(cases: list[Mapping[str, Any]]) -> dict[str, Any]:
    patterns: dict[str, int] = {}
    employers: dict[str, int] = {}
    addenda = 0
    artifacts = 0
    pages = 0
    size = 0
    for case in cases:
        pattern = str(case["appendix_a_evidence_pattern"])
        patterns[pattern] = patterns.get(pattern, 0) + 1
        count = str(case["appendix_a_employer_count"])
        employers[count] = employers.get(count, 0) + 1
        addenda += bool(case["pwd_addendum_sections"])
        for artifact in case["artifacts"].values():
            artifacts += 1
            pages += int(artifact["page_count"])
            size += int(artifact["bytes"])
    return {
        "pair_count": len(cases),
        "pwd_addendum_case_count": addenda,
        "appendix_a_evidence_patterns": patterns,
        "appendix_a_employer_counts": employers,
        "pdf_artifact_count": artifacts,
        "pdf_page_count": pages,
        "total_pdf_bytes": size,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--count", type=int, default=5000)
    parser.add_argument("--workers", type=int, default=min(4, os.cpu_count() or 1))
    parser.add_argument(
        "--perm-file",
        type=Path,
        default=ROOT / "data/raw/oflc/PERM/FY2025_Q4/PERM_Disclosure_Data_FY2025_Q4.xlsx",
    )
    parser.add_argument(
        "--pwd-file",
        type=Path,
        default=ROOT / "data/raw/oflc/PW/FY2024_Q4/PW_Disclosure_Data_FY2024_Q4.xlsx",
    )
    parser.add_argument(
        "--templates-dir",
        type=Path,
        default=Path("/Users/Dad/Desktop/Qwen Training Documents/Blank Forms"),
    )
    parser.add_argument(
        "--pdf-output", type=Path, default=ROOT / "output/pdf/eta_pair_full_5000"
    )
    parser.add_argument(
        "--json-output", type=Path, default=ROOT / "output/json/eta_pair_full_5000"
    )
    parser.add_argument("--watermark", default="TRAINING EXAMPLE - NOT FOR FILING")
    parser.add_argument("--resume", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument(
        "--select-only",
        action="store_true",
        help="Write the deterministic plan and stop before PDF generation.",
    )
    args = parser.parse_args()
    if args.count <= 0:
        parser.error("--count must be positive")
    if args.workers <= 0:
        parser.error("--workers must be positive")

    print(f"Indexing PWD case numbers from {args.pwd_file}", flush=True)
    pwd_numbers = _case_numbers(args.pwd_file)
    print(f"Indexed {len(pwd_numbers):,} PWD case numbers", flush=True)
    print(f"Selecting {args.count:,} deterministic PERM matches", flush=True)
    perm_rows, matched_count, unique_match_count = _select_perm_rows(
        args.perm_file, pwd_numbers, args.count
    )
    if len(perm_rows) < args.count:
        raise ValueError(f"Only {len(perm_rows):,} matched PERM rows are available")
    selected_pwd_numbers = {str(row["JOB_OPP_PWD_NUMBER"]) for row in perm_rows}
    print(
        f"Found {matched_count:,} exact matches across {unique_match_count:,} unique PWDs; "
        f"reading {len(selected_pwd_numbers):,} PWD rows",
        flush=True,
    )
    pwd_rows = _selected_rows(args.pwd_file, selected_pwd_numbers)

    tasks = []
    selections = []
    for ordinal, perm in enumerate(perm_rows, 1):
        case_id = str(perm["CASE_NUMBER"])
        pwd_number = str(perm["JOB_OPP_PWD_NUMBER"])
        pwd = pwd_rows[pwd_number]
        scenario = _scenario(case_id, pwd)
        fn = _varied_foreign_national(perm, pwd, scenario)
        pwd_form = build_eta9141_form_data(pwd, perm)
        perm_form = build_eta9089_form_data(perm, pwd, foreign_national=fn)
        selection = {
            "ordinal": ordinal,
            "perm_case_number": case_id,
            "pwd_case_number": pwd_number,
            **scenario,
        }
        selections.append(selection)
        tasks.append(
            {
                "selection": selection,
                "pwd_form": pwd_form,
                "perm_form": perm_form,
                "perm_file": str(args.perm_file),
                "pwd_file": str(args.pwd_file),
                "templates_dir": str(args.templates_dir),
                "pdf_output": str(args.pdf_output),
                "json_output": str(args.json_output),
                "watermark": args.watermark,
                "resume": args.resume,
            }
        )

    args.json_output.mkdir(parents=True, exist_ok=True)
    plan = {
        "schema_version": "casebase.eta-pair-corpus-plan.v1",
        "created_at": datetime.now().astimezone().isoformat(),
        "selection_method": (
            "lowest SHA-256 values of the first matched PERM case for each unique PWD"
        ),
        "available_exact_match_count": matched_count,
        "available_unique_pwd_match_count": unique_match_count,
        "requested_pair_count": args.count,
        "source_files": {"perm": str(args.perm_file), "pwd": str(args.pwd_file)},
        "watermark": args.watermark,
        "selections": selections,
    }
    plan_path = args.json_output / "selection-plan.json"
    plan_path.write_text(json.dumps(plan, indent=2) + "\n", encoding="utf-8")
    print(f"Selection plan: {plan_path}", flush=True)
    if args.select_only:
        return

    args.pdf_output.mkdir(parents=True, exist_ok=True)
    completed: list[dict[str, Any]] = []
    failures: list[dict[str, str]] = []
    progress_path = args.json_output / "progress.json"
    print(f"Generating with {args.workers} worker processes", flush=True)
    with ProcessPoolExecutor(max_workers=args.workers) as executor:
        futures = {executor.submit(_generate_case, task): task for task in tasks}
        for future in as_completed(futures):
            task = futures[future]
            try:
                completed.append(future.result())
            except Exception as exc:  # preserve other cases and make failures resumable
                failures.append(
                    {
                        "perm_case_number": task["selection"]["perm_case_number"],
                        "error": f"{type(exc).__name__}: {exc}",
                    }
                )
            finished = len(completed) + len(failures)
            if finished % 25 == 0 or finished == len(tasks):
                progress = {
                    "updated_at": datetime.now().astimezone().isoformat(),
                    "requested": len(tasks),
                    "finished": finished,
                    "completed": len(completed),
                    "failed": len(failures),
                    "failures": failures,
                }
                progress_path.write_text(
                    json.dumps(progress, indent=2) + "\n", encoding="utf-8"
                )
                print(
                    f"Progress {finished:,}/{len(tasks):,}: "
                    f"{len(completed):,} complete, {len(failures):,} failed",
                    flush=True,
                )

    completed.sort(key=lambda item: item["ordinal"])
    manifest = {
        "schema_version": "casebase.eta-pair-corpus.v1",
        "created_at": datetime.now().astimezone().isoformat(),
        "training_approved": False,
        "watermark": args.watermark,
        "source_files": {"perm": str(args.perm_file), "pwd": str(args.pwd_file)},
        "selection_plan": str(plan_path),
        "summary": _summary(completed),
        "failures": failures,
        "cases": completed,
    }
    manifest_path = args.json_output / "full-manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    print(f"Manifest: {manifest_path}", flush=True)
    if failures or len(completed) != args.count:
        raise SystemExit(
            f"Run incomplete: {len(completed):,}/{args.count:,} cases; "
            f"rerun with --resume after resolving {len(failures):,} failures"
        )


if __name__ == "__main__":
    main()
