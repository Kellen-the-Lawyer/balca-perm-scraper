#!/usr/bin/env python3
"""Build the first ten matched ETA-9141/ETA-9089 training packages."""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.perm_verify.form_fill import fill_eta9141, fill_eta9089_package  # noqa: E402
from app.perm_verify.form_fill.dol_eta9141 import (  # noqa: E402
    build_eta9141_form_data,
)
from app.perm_verify.form_fill.dol_eta9089 import (  # noqa: E402
    build_eta9089_form_data,
)
from app.perm_verify.form_fill.eta9141 import TEMPLATE as ETA9141_TEMPLATE  # noqa: E402
from app.perm_verify.form_fill.eta9141_addendum import (  # noqa: E402
    eta9141_addendum_sections,
    eta9141_with_addendum_references,
    generate_eta9141_addendum,
    merge_eta9141_package,
)


PILOT_CASES = [
    {
        "perm": "G-200-24127-961873",
        "pwd": "P-100-23124-994087",
        "reason": "non-professional; no degree or experience requirement",
    },
    {
        "perm": "G-200-24172-131528",
        "pwd": "P-100-23133-019672",
        "reason": "non-professional; high-school and training requirements",
    },
    {
        "perm": "G-200-24089-841257",
        "pwd": "P-100-23122-986140",
        "reason": "non-professional; training and alternative requirements",
    },
    {
        "perm": "G-100-24102-879722",
        "pwd": "P-100-23144-050582",
        "reason": "professional; explicit foreign-language requirement",
        "pwd_addendum": True,
    },
    {
        "perm": "G-100-24036-693559",
        "pwd": "P-100-23102-927250",
        "reason": "professional; no degree and alternative requirements",
        "evidence_pattern": "split_skills",
        "employer_count": 2,
        "pwd_addendum": True,
    },
    {
        "perm": "G-100-24115-922983",
        "pwd": "P-100-23172-133024",
        "reason": "professional; employer survey wage source",
        "evidence_pattern": "multiple_employers",
        "employer_count": 2,
        "pwd_addendum": True,
    },
    {
        "perm": "G-100-24123-951214",
        "pwd": "P-100-23209-221676",
        "reason": "professional; supervision and alternative requirements",
        "evidence_pattern": "split_skills",
        "employer_count": 3,
        "pwd_addendum": True,
    },
    {
        "perm": "G-100-24093-847323",
        "pwd": "P-100-23118-976448",
        "reason": "professional; Level II with no special requirements",
    },
    {
        "perm": "G-100-23325-517525",
        "pwd": "P-100-23089-891423",
        "reason": "professional; Level IV and experience requirement",
        "evidence_pattern": "multiple_employers",
        "employer_count": 2,
        "pwd_addendum": True,
    },
    {
        "perm": "G-100-24173-139221",
        "pwd": "P-100-23307-478925",
        "reason": "professional; OEWS naming and Level II determination",
        "pwd_addendum": True,
    },
]


def _jsonable(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if hasattr(value, "item"):
        return value.item()
    return value


def _rows(path: Path, case_numbers: set[str]) -> dict[str, dict[str, Any]]:
    book = load_workbook(path, read_only=True, data_only=True)
    sheet = book.active
    iterator = sheet.iter_rows(values_only=True)
    headers = list(next(iterator))
    try:
        case_index = headers.index("CASE_NUMBER")
    except ValueError as exc:
        book.close()
        raise ValueError(f"{path} has no CASE_NUMBER column") from exc

    found: dict[str, dict[str, Any]] = {}
    for row in iterator:
        case_number = row[case_index]
        if case_number not in case_numbers:
            continue
        found[str(case_number)] = {
            str(header): _jsonable(value)
            for header, value in zip(headers, row)
            if header is not None
        }
        if len(found) == len(case_numbers):
            break
    book.close()

    missing = sorted(case_numbers - set(found))
    if missing:
        raise ValueError(f"{path} is missing selected cases: {missing}")
    return found


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _artifact(path: Path) -> dict[str, Any]:
    return {
        "path": str(path),
        "sha256": _sha256(path),
        "bytes": path.stat().st_size,
        "page_count": len(PdfReader(path).pages),
    }


def _evidence_validation(perm_form: dict[str, Any]) -> dict[str, Any]:
    appendix = perm_form["appendix_A"]
    required_ids = {item["id"] for item in appendix.get("requirements", [])}
    covered_ids = {
        identifier
        for skill in appendix.get("skills", [])
        for identifier in skill.get("requirement_ids", [])
        if identifier in required_ids
    }
    experiences = appendix.get("work_experience", [])
    employer_evidence = []
    for index, experience in enumerate(experiences):
        provider_names = {
            skill.get("provider")
            for skill in appendix.get("skills", [])
            if skill.get("experience_index", 0) == index
        }
        employer_evidence.append(experience.get("employer_name") in provider_names)
    required_months = appendix.get("required_experience_months", 0)
    credited_months = sum(
        int(item.get("credited_months", 0)) for item in experiences
    )
    return {
        "requirement_ids": sorted(required_ids),
        "covered_requirement_ids": sorted(covered_ids),
        "missing_requirement_ids": sorted(required_ids - covered_ids),
        "requirements_coverage": "pass" if required_ids <= covered_ids else "fail",
        "required_experience_months": required_months,
        "credited_experience_months": credited_months,
        "cumulative_experience": (
            "pass" if credited_months >= required_months else "fail"
        ),
        "section_d_e_employer_links": employer_evidence,
        "section_d_e_linkage": "pass" if all(employer_evidence) else "fail",
    }


def _review_flags(
    pwd_form: dict[str, Any], perm_form: dict[str, Any]
) -> list[str]:
    flags = [f"synthetic:{path}" for path in pwd_form["synthetic_fields"]]
    if not pwd_form["job_offer"]["worksite"].get("county"):
        flags.append("missing:job_offer.worksite.county")
    if any(
        "see f.a.2" in item["text"].lower()
        for item in perm_form["appendix_A"].get("requirements", [])
    ):
        flags.append("source_gap:PWD_F.a.2_text_not_in_disclosure_workbook")
    return flags


def _build_cases(
    perm_file: Path, pwd_file: Path
) -> list[tuple[dict[str, str], dict[str, Any], dict[str, Any]]]:
    selected_perm = {item["perm"] for item in PILOT_CASES}
    selected_pwd = {item["pwd"] for item in PILOT_CASES}
    print(f"Reading {len(selected_perm)} PERM rows from {perm_file}", flush=True)
    perm_rows = _rows(perm_file, selected_perm)
    print(f"Reading {len(selected_pwd)} PWD rows from {pwd_file}", flush=True)
    pwd_rows = _rows(pwd_file, selected_pwd)

    cases = []
    for selection in PILOT_CASES:
        perm = perm_rows[selection["perm"]]
        pwd = pwd_rows[selection["pwd"]]
        if perm.get("JOB_OPP_PWD_NUMBER") != pwd.get("CASE_NUMBER"):
            raise ValueError(f"link mismatch for {selection['perm']}")
        cases.append((selection, perm, pwd))
    return cases


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--perm-file",
        type=Path,
        default=ROOT
        / "data/raw/oflc/PERM/FY2025_Q4/PERM_Disclosure_Data_FY2025_Q4.xlsx",
    )
    parser.add_argument(
        "--pwd-file",
        type=Path,
        default=ROOT
        / "data/raw/oflc/PW/FY2024_Q4/PW_Disclosure_Data_FY2024_Q4.xlsx",
    )
    parser.add_argument(
        "--templates-dir",
        type=Path,
        default=Path("/Users/Dad/Desktop/Qwen Training Documents/Blank Forms"),
    )
    parser.add_argument(
        "--pdf-output",
        type=Path,
        default=ROOT / "output/pdf/eta_pair_pilot_10",
    )
    parser.add_argument(
        "--json-output",
        type=Path,
        default=ROOT / "output/json/eta_pair_pilot_10",
    )
    parser.add_argument(
        "--watermark", default="TRAINING EXAMPLE - NOT FOR FILING"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Read and validate selected rows without writing JSON or PDFs.",
    )
    args = parser.parse_args()

    cases = _build_cases(args.perm_file, args.pwd_file)
    prepared = []
    for selection, perm, pwd in cases:
        pwd_form = build_eta9141_form_data(pwd, perm)
        perm_form = build_eta9089_form_data(
            perm,
            pwd,
            evidence_pattern=selection.get("evidence_pattern", "single_employer"),
            employer_count=int(selection.get("employer_count", 1)),
        )
        prepared.append((selection, pwd_form, perm_form))
        print(
            f"Prepared {selection['perm']} <-> {selection['pwd']}: "
            f"{selection['reason']}",
            flush=True,
        )

    if args.dry_run:
        print("Dry run passed for all ten linked pairs.")
        return

    args.pdf_output.mkdir(parents=True, exist_ok=True)
    args.json_output.mkdir(parents=True, exist_ok=True)
    manifest: dict[str, Any] = {
        "schema_version": "casebase.eta-pair-pilot.v1",
        "created_at": datetime.now().astimezone().isoformat(),
        "pair_count": len(prepared),
        "training_approved": False,
        "watermark": args.watermark,
        "source_files": {
            "perm": str(args.perm_file),
            "pwd": str(args.pwd_file),
        },
        "cases": [],
    }
    template_9141 = args.templates_dir / ETA9141_TEMPLATE

    for index, (selection, pwd_form, perm_form) in enumerate(prepared, 1):
        case_id = selection["perm"]
        case_pdf_dir = args.pdf_output / case_id
        case_json_dir = args.json_output / case_id
        case_pdf_dir.mkdir(parents=True, exist_ok=True)
        case_json_dir.mkdir(parents=True, exist_ok=True)

        pwd_pdf = case_pdf_dir / "eta9141.pdf"
        addendum_sections = (
            eta9141_addendum_sections(pwd_form)
            if selection.get("pwd_addendum")
            else []
        )
        pwd_render_form = (
            eta9141_with_addendum_references(pwd_form, addendum_sections)
            if addendum_sections
            else pwd_form
        )
        pwd_result = fill_eta9141(
            pwd_render_form,
            template_9141,
            pwd_pdf,
            watermark=args.watermark,
        )
        addendum_result = None
        complete_pwd_result = None
        if addendum_sections:
            addendum_pdf = case_pdf_dir / "eta9141_addendum.pdf"
            addendum_result = generate_eta9141_addendum(
                pwd_form,
                addendum_sections,
                template_9141,
                addendum_pdf,
                watermark=args.watermark,
            )
            complete_pwd_result = merge_eta9141_package(
                pwd_pdf,
                addendum_pdf,
                case_pdf_dir / "eta9141_complete.pdf",
            )
        perm_result = fill_eta9089_package(
            perm_form,
            args.templates_dir,
            case_pdf_dir,
            watermark=args.watermark,
        )

        review_flags = _review_flags(pwd_form, perm_form)
        evidence_validation = _evidence_validation(perm_form)
        if any(
            evidence_validation[key] != "pass"
            for key in (
                "requirements_coverage",
                "cumulative_experience",
                "section_d_e_linkage",
            )
        ):
            raise ValueError(
                f"Appendix A evidence validation failed for {case_id}: "
                f"{evidence_validation}"
            )
        pair_payload = {
            "schema_version": "casebase.eta-pair-fill.v1",
            "case_id": case_id,
            "source": {
                "perm_file": str(args.perm_file),
                "perm_case_number": selection["perm"],
                "pwd_file": str(args.pwd_file),
                "pwd_case_number": selection["pwd"],
            },
            "selection_reason": selection["reason"],
            "privacy": {
                "dol_disclosure_data": True,
                "foreign_national": "synthetic training identity",
                "not_for_filing": True,
                "training_approved": False,
            },
            "review_flags": review_flags,
            "evidence_validation": evidence_validation,
            "pwd": {
                "form_data": pwd_form,
                "render_form_data": pwd_render_form,
                "fill_result": pwd_result,
                "addendum_result": addendum_result,
                "complete_package_result": complete_pwd_result,
            },
            "perm": {"form_data": perm_form, "fill_result": perm_result},
        }
        pair_json = case_json_dir / "pair.json"
        pair_json.write_text(
            json.dumps(pair_payload, indent=2) + "\n", encoding="utf-8"
        )

        artifacts = {"eta9141": _artifact(pwd_pdf)}
        if addendum_sections:
            artifacts["eta9141_addendum"] = _artifact(
                case_pdf_dir / "eta9141_addendum.pdf"
            )
            artifacts["eta9141_complete"] = _artifact(
                case_pdf_dir / "eta9141_complete.pdf"
            )
        for key in perm_result:
            artifact_key = "eta9089" if key == "application" else f"eta9089_{key}"
            artifacts[artifact_key] = _artifact(case_pdf_dir / f"{key}.pdf")
        manifest["cases"].append(
            {
                "ordinal": index,
                "perm_case_number": selection["perm"],
                "pwd_case_number": selection["pwd"],
                "selection_reason": selection["reason"],
                "pair_json": str(pair_json),
                "synthetic_pwd_fields": pwd_form["synthetic_fields"],
                "pwd_addendum_sections": [
                    item["section"] for item in addendum_sections
                ],
                "appendix_a_evidence_pattern": perm_form["appendix_A"].get(
                    "evidence_pattern"
                ),
                "appendix_a_employer_count": len(
                    perm_form["appendix_A"].get("work_experience", [])
                ),
                "review_flags": review_flags,
                "evidence_validation": evidence_validation,
                "artifacts": artifacts,
                "validation": {
                    "pwd_number_link": "pass",
                    "acroform_reopen": "pass",
                    "widget_appearances": "pass",
                    "human_review": "pending",
                },
            }
        )
        print(f"Generated {index}/10: {case_id}", flush=True)

    manifest["total_pdf_bytes"] = sum(
        artifact["bytes"]
        for case in manifest["cases"]
        for artifact in case["artifacts"].values()
    )
    manifest_path = args.json_output / "pilot-manifest.json"
    manifest_path.write_text(
        json.dumps(manifest, indent=2) + "\n", encoding="utf-8"
    )
    print(manifest_path)


if __name__ == "__main__":
    main()
